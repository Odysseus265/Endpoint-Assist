"""
Endpoint Assist - Excel Report Generator
Generates Excel spreadsheet reports for system diagnostics
"""

import io
from datetime import datetime
import psutil
import platform
import socket

# Try to import openpyxl, provide fallback if not available
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, BarChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class ExcelReportGenerator:
    """Generate Excel reports for system diagnostics"""
    
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel reports. Install with: pip install openpyxl")
        
        # Define styles
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        self.title_font = Font(bold=True, color="d00000", size=16)
        self.subtitle_font = Font(bold=True, size=12)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.warning_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        self.danger_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        self.success_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    
    def _style_header_row(self, ws, row, columns):
        """Apply header styling to a row"""
        for col in range(1, columns + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_align
    
    def _auto_width(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_title(self, ws, title, row=1):
        """Add a title to the worksheet"""
        ws.cell(row=row, column=1, value=title).font = self.title_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    
    def _get_status_fill(self, value, warning_threshold, danger_threshold):
        """Get fill color based on value thresholds"""
        if value >= danger_threshold:
            return self.danger_fill
        elif value >= warning_threshold:
            return self.warning_fill
        return self.success_fill
    
    def create_system_sheet(self, wb):
        """Create system information sheet"""
        ws = wb.active
        ws.title = "System Info"
        
        # Title
        self._add_title(ws, "🛡️ Endpoint Assist - System Report")
        ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(row=3, column=1, value=f"Hostname: {socket.gethostname()}")
        
        # OS Information
        ws.cell(row=5, column=1, value="Operating System Information").font = self.subtitle_font
        
        os_data = [
            ["Property", "Value"],
            ["Operating System", f"{platform.system()} {platform.release()}"],
            ["Version", platform.version()],
            ["Machine", platform.machine()],
            ["Processor", platform.processor()],
            ["Hostname", socket.gethostname()],
            ["Boot Time", datetime.fromtimestamp(psutil.boot_time()).strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        for i, row_data in enumerate(os_data, start=6):
            for j, value in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=j, value=value)
                cell.border = self.border
                if i == 6:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
        
        self._auto_width(ws)
        return ws
    
    def create_performance_sheet(self, wb):
        """Create performance metrics sheet"""
        ws = wb.create_sheet("Performance")
        
        self._add_title(ws, "📊 Performance Metrics")
        
        # CPU Info
        cpu_percent = psutil.cpu_percent(interval=1)
        memory = psutil.virtual_memory()
        
        perf_data = [
            ["Metric", "Value", "Status"],
            ["CPU Usage", f"{cpu_percent}%", "Critical" if cpu_percent > 80 else "Warning" if cpu_percent > 60 else "Good"],
            ["CPU Cores (Physical)", psutil.cpu_count(logical=False), "Info"],
            ["CPU Cores (Logical)", psutil.cpu_count(logical=True), "Info"],
            ["Memory Total", f"{round(memory.total / (1024**3), 2)} GB", "Info"],
            ["Memory Used", f"{round(memory.used / (1024**3), 2)} GB", "Info"],
            ["Memory Available", f"{round(memory.available / (1024**3), 2)} GB", "Info"],
            ["Memory Usage", f"{memory.percent}%", "Critical" if memory.percent > 85 else "Warning" if memory.percent > 70 else "Good"],
        ]
        
        for i, row_data in enumerate(perf_data, start=3):
            for j, value in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=j, value=value)
                cell.border = self.border
                if i == 3:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                elif j == 3:  # Status column
                    if value == "Critical":
                        cell.fill = self.danger_fill
                    elif value == "Warning":
                        cell.fill = self.warning_fill
                    elif value == "Good":
                        cell.fill = self.success_fill
        
        self._auto_width(ws)
        return ws
    
    def create_disk_sheet(self, wb):
        """Create disk information sheet"""
        ws = wb.create_sheet("Disks")
        
        self._add_title(ws, "💾 Disk Information")
        
        headers = ["Drive", "File System", "Total (GB)", "Used (GB)", "Free (GB)", "Usage %", "Status"]
        for j, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=j, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        row = 4
        for partition in psutil.disk_partitions():
            try:
                usage = psutil.disk_usage(partition.mountpoint)
                status = "Critical" if usage.percent > 90 else "Warning" if usage.percent > 75 else "Good"
                
                data = [
                    partition.device,
                    partition.fstype,
                    round(usage.total / (1024**3), 2),
                    round(usage.used / (1024**3), 2),
                    round(usage.free / (1024**3), 2),
                    f"{usage.percent}%",
                    status
                ]
                
                for j, value in enumerate(data, start=1):
                    cell = ws.cell(row=row, column=j, value=value)
                    cell.border = self.border
                    if j == 7:  # Status column
                        if value == "Critical":
                            cell.fill = self.danger_fill
                        elif value == "Warning":
                            cell.fill = self.warning_fill
                        else:
                            cell.fill = self.success_fill
                
                row += 1
            except:
                pass
        
        self._auto_width(ws)
        return ws
    
    def create_network_sheet(self, wb):
        """Create network information sheet"""
        ws = wb.create_sheet("Network")
        
        self._add_title(ws, "🌐 Network Information")
        
        # Get local IP
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(('8.8.8.8', 80))
            local_ip = s.getsockname()[0]
            s.close()
        except:
            local_ip = '127.0.0.1'
        
        ws.cell(row=3, column=1, value="Local IP:").font = Font(bold=True)
        ws.cell(row=3, column=2, value=local_ip)
        
        # Network interfaces
        ws.cell(row=5, column=1, value="Network Interfaces").font = self.subtitle_font
        
        headers = ["Interface", "IP Address", "Netmask"]
        for j, header in enumerate(headers, start=1):
            cell = ws.cell(row=6, column=j, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        row = 7
        for name, addrs in psutil.net_if_addrs().items():
            for addr in addrs:
                if addr.family == socket.AF_INET:
                    data = [name, addr.address, addr.netmask or "N/A"]
                    for j, value in enumerate(data, start=1):
                        cell = ws.cell(row=row, column=j, value=value)
                        cell.border = self.border
                    row += 1
        
        # Network stats
        ws.cell(row=row + 2, column=1, value="Network Statistics").font = self.subtitle_font
        net_io = psutil.net_io_counters()
        
        stats_data = [
            ["Metric", "Value"],
            ["Bytes Sent", f"{round(net_io.bytes_sent / (1024**2), 2)} MB"],
            ["Bytes Received", f"{round(net_io.bytes_recv / (1024**2), 2)} MB"],
            ["Packets Sent", net_io.packets_sent],
            ["Packets Received", net_io.packets_recv],
        ]
        
        start_row = row + 3
        for i, row_data in enumerate(stats_data):
            for j, value in enumerate(row_data, start=1):
                cell = ws.cell(row=start_row + i, column=j, value=value)
                cell.border = self.border
                if i == 0:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
        
        self._auto_width(ws)
        return ws
    
    def create_processes_sheet(self, wb):
        """Create processes sheet"""
        ws = wb.create_sheet("Processes")
        
        self._add_title(ws, "⚡ Top Processes")
        
        headers = ["PID", "Process Name", "CPU %", "Memory %", "Status"]
        for j, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=j, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        processes = []
        for proc in psutil.process_iter(['pid', 'name', 'cpu_percent', 'memory_percent', 'status']):
            try:
                pinfo = proc.info
                if pinfo['cpu_percent'] is not None:
                    processes.append(pinfo)
            except:
                pass
        
        # Sort by CPU
        processes.sort(key=lambda x: x['cpu_percent'] or 0, reverse=True)
        
        for i, proc in enumerate(processes[:25], start=4):  # Top 25 processes
            data = [
                proc['pid'],
                proc['name'][:40],
                f"{round(proc['cpu_percent'], 1)}%",
                f"{round(proc['memory_percent'], 1)}%" if proc['memory_percent'] else "N/A",
                proc['status']
            ]
            for j, value in enumerate(data, start=1):
                cell = ws.cell(row=i, column=j, value=value)
                cell.border = self.border
        
        self._auto_width(ws)
        return ws
    
    def generate_full_report(self):
        """Generate a comprehensive Excel report"""
        wb = Workbook()
        
        self.create_system_sheet(wb)
        self.create_performance_sheet(wb)
        self.create_disk_sheet(wb)
        self.create_network_sheet(wb)
        self.create_processes_sheet(wb)
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def generate_system_report(self):
        """Generate system-only Excel report"""
        wb = Workbook()
        self.create_system_sheet(wb)
        self.create_performance_sheet(wb)
        self.create_disk_sheet(wb)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def generate_network_report(self):
        """Generate network-only Excel report"""
        wb = Workbook()
        
        # Rename default sheet
        ws = wb.active
        ws.title = "Network"
        self._add_title(ws, "🌐 Network Report")
        
        # Add network info directly to first sheet
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(('8.8.8.8', 80))
            local_ip = s.getsockname()[0]
            s.close()
        except:
            local_ip = '127.0.0.1'
        
        ws.cell(row=3, column=1, value="Local IP:").font = Font(bold=True)
        ws.cell(row=3, column=2, value=local_ip)
        ws.cell(row=4, column=1, value="Hostname:").font = Font(bold=True)
        ws.cell(row=4, column=2, value=socket.gethostname())
        
        # Interfaces
        headers = ["Interface", "IP Address", "Netmask"]
        for j, header in enumerate(headers, start=1):
            cell = ws.cell(row=6, column=j, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        row = 7
        for name, addrs in psutil.net_if_addrs().items():
            for addr in addrs:
                if addr.family == socket.AF_INET:
                    for j, value in enumerate([name, addr.address, addr.netmask or "N/A"], start=1):
                        cell = ws.cell(row=row, column=j, value=value)
                        cell.border = self.border
                    row += 1
        
        self._auto_width(ws)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer


# Helper functions
def generate_excel_report(report_type='full'):
    """Generate Excel report of specified type"""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")
    
    generator = ExcelReportGenerator()
    
    if report_type == 'system':
        return generator.generate_system_report()
    elif report_type == 'network':
        return generator.generate_network_report()
    else:
        return generator.generate_full_report()
